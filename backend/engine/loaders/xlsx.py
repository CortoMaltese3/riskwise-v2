"""XLSX entity loader — riskwise-side replacement for ``Entity.from_excel``.

Reads the workbook structure produced by CLIMADA's entity XLSX format and
returns an :class:`backend.engine.types.EntityBundle`. The two paths run
side by side for the duration of Phase 6 so parity tests can pin the new
path against CLIMADA on the same fixtures.

Workbook layout expected
------------------------
* ``assets``            — exposure rows (value, lat, lon, impf_id, …).
* ``impact_functions``  — impact-function curve rows.
* ``measures``          — adaptation measures.
* ``discount``          — per-year discount rates.
* ``names`` (optional)  — metadata; ``reference_year`` row → ``ref_year``.

discount_rate is the arithmetic mean of all per-year rates found in the
``discount`` sheet ``discount_rate`` column. CLIMADA's ``DiscRates`` class
exposes the raw per-year table; a single scalar representative is needed
for ``EntityBundle.discount_rate``. The mean is stable and symmetric — a
caller that needs a different scalar (e.g. the final-year rate) can
re-derive it from the CLIMADA path.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, overload

import numpy as np
import openpyxl

from backend.engine.loaders._errors import EntityLoadError
from backend.engine.types import EntityBundle, ExposureArrays, ImpactFunctionSpec, MeasureSpec

__all__ = ["load_entity_xlsx"]


def load_entity_xlsx(path: Path) -> EntityBundle:
    """Load a CLIMADA-format entity XLSX file into :class:`EntityBundle`.

    Parameters
    ----------
    path:
        Path to the ``.xlsx`` file written by CLIMADA's entity export or
        following the riskwise entity workbook template.

    Returns
    -------
    EntityBundle
        Frozen aggregator with exposures, impact-function specs, measures,
        a representative discount rate, and a reference year.

    Raises
    ------
    EntityLoadError
        File missing, required sheet or column missing, invalid cell
        value, or impact-function validation failure (non-monotonic curve,
        duplicate id, inconsistent units). The message names the offending
        row or column.
    """
    path = Path(path)
    if not path.is_file():
        raise EntityLoadError(f"Entity file not found: {path}")

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except (OSError, ValueError, KeyError) as exc:
        # openpyxl wraps the zip/xml parse errors in InvalidFileException
        # (a ValueError) and surfaces missing-file as OSError.
        raise EntityLoadError(f"Cannot open entity file {path}: {exc}") from exc

    try:
        exposures = _parse_assets(wb, path)
        specs = _parse_impact_functions(wb, path)
        measures = _parse_measures(wb, path)
        discount_rate = _parse_discount_rate(wb, path)
        ref_year = _parse_ref_year(wb, path)
    finally:
        wb.close()

    _validate_specs(specs, path)

    return EntityBundle(
        exposures=exposures,
        impfset_specs=specs,
        measures=measures,
        discount_rate=discount_rate,
        ref_year=ref_year,
    )


# ---------------------------------------------------------------------------
# Sheet parsers
# ---------------------------------------------------------------------------


def _parse_assets(wb: openpyxl.Workbook, path: Path) -> ExposureArrays:
    _require_sheet(wb, "assets", path)
    ws = wb["assets"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise EntityLoadError(f"Empty 'assets' sheet in {path.name}")

    header = _normalise_header(rows[0])
    lat_c = _require_col(header, "latitude", "assets", path)
    lon_c = _require_col(header, "longitude", "assets", path)
    val_c = _require_col(header, "value", "assets", path)
    # CLIMADA uses either generic 'impf_' or peril-specific 'impf_{HAZ_TYPE}'.
    impf_c = _find_col_prefix(header, "impf_")
    if impf_c is None:
        raise EntityLoadError(
            f"Missing required column 'impf_' (or 'impf_<haz_type>') "
            f"in sheet 'assets' of {path.name}"
        )
    ded_c = _find_col(header, "deductible")
    cov_c = _find_col(header, "cover")
    unit_c = _find_col(header, "value_unit")

    data = [r for r in rows[1:] if r[val_c] is not None]
    if not data:
        raise EntityLoadError(f"No data rows in 'assets' sheet of {path.name}")

    n = len(data)
    try:
        values = np.array([float(r[val_c]) for r in data], dtype=np.float64)
        lat = np.array([float(r[lat_c]) for r in data], dtype=np.float64)
        lon = np.array([float(r[lon_c]) for r in data], dtype=np.float64)
        impf_id = np.array([int(r[impf_c]) for r in data], dtype=np.int64)
    except (TypeError, ValueError) as exc:
        raise EntityLoadError(
            f"Invalid numeric value in 'assets' sheet of {path.name}: {exc}"
        ) from exc

    centroid_idx = np.arange(n, dtype=np.int64)
    deductible = _opt_array(data, ded_c, "deductible", path)
    cover = _opt_array(data, cov_c, "cover", path)

    value_unit = "USD"
    if unit_c is not None:
        units = {r[unit_c] for r in data if r[unit_c] is not None}
        if units:
            value_unit = str(next(iter(units)))

    return ExposureArrays(
        values=values,
        centroid_idx=centroid_idx,
        impf_id=impf_id,
        lat=lat,
        lon=lon,
        deductible=deductible,
        cover=cover,
        value_unit=value_unit,
    )


def _parse_impact_functions(wb: openpyxl.Workbook, path: Path) -> list[ImpactFunctionSpec]:
    _require_sheet(wb, "impact_functions", path)
    ws = wb["impact_functions"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = _normalise_header(rows[0])
    id_c = _require_col(header, "impact_fun_id", "impact_functions", path)
    int_c = _require_col(header, "intensity", "impact_functions", path)
    mdd_c = _require_col(header, "mdd", "impact_functions", path)
    paa_c = _require_col(header, "paa", "impact_functions", path)
    haz_c = _require_col(header, "peril_id", "impact_functions", path)
    unit_c = _require_col(header, "intensity_unit", "impact_functions", path)
    name_c = _require_col(header, "name", "impact_functions", path)

    # Group curve rows by (fun_id, haz_type, name, intensity_unit).
    groups: dict[tuple[int, str, str, str], list[tuple[float, float, float]]] = {}
    for row_num, row in enumerate(rows[1:], start=2):
        if row[id_c] is None:
            continue
        try:
            fun_id = int(row[id_c])
            haz_type = str(row[haz_c]).strip()
            name = str(row[name_c]).strip()
            unit = str(row[unit_c]).strip()
            intensity = float(row[int_c])
            mdd = float(row[mdd_c])
            paa = float(row[paa_c])
        except (TypeError, ValueError) as exc:
            raise EntityLoadError(
                f"Invalid value in 'impact_functions' sheet row {row_num} of {path.name}: {exc}"
            ) from exc
        groups.setdefault((fun_id, haz_type, name, unit), []).append((intensity, mdd, paa))

    specs: list[ImpactFunctionSpec] = []
    for (fun_id, haz_type, name, unit), curve in groups.items():
        intensities, mdds, paas = zip(*curve, strict=False)
        specs.append(
            ImpactFunctionSpec(
                haz_type=haz_type,
                exp_type=name,
                id=fun_id,
                name=name,
                intensity_unit=unit,
                intensity=tuple(intensities),
                mdd=tuple(mdds),
                paa=tuple(paas),
            )
        )

    return specs


def _parse_measures(wb: openpyxl.Workbook, path: Path) -> list[MeasureSpec]:
    _require_sheet(wb, "measures", path)
    ws = wb["measures"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = _normalise_header(rows[0])
    name_c = _require_col(header, "name", "measures", path)
    haz_c = _require_col(header, "peril_id", "measures", path)
    cost_c = _find_col(header, "cost")
    freq_c = _find_col(header, "hazard high frequency cutoff")
    inten_a_c = _find_col(header, "hazard intensity impact a")
    mdd_a_c = _find_col(header, "mdd impact a")
    mdd_b_c = _find_col(header, "mdd impact b")
    paa_a_c = _find_col(header, "paa impact a")
    paa_b_c = _find_col(header, "paa impact b")

    measures: list[MeasureSpec] = []
    for row_num, row in enumerate(rows[1:], start=2):
        if row[name_c] is None:
            continue
        try:
            name = str(row[name_c]).strip()
            haz_type = str(row[haz_c]).strip() if row[haz_c] is not None else ""
            cost = _opt_float(row, cost_c, 0.0)
            freq_cutoff = _opt_float(row, freq_c)
            inten_imp = _opt_float(row, inten_a_c)
            mdd_a = _opt_float(row, mdd_a_c)
            mdd_b = _opt_float(row, mdd_b_c)
            paa_a = _opt_float(row, paa_a_c)
            paa_b = _opt_float(row, paa_b_c)
        except (TypeError, ValueError) as exc:
            raise EntityLoadError(
                f"Invalid value in 'measures' sheet row {row_num} of {path.name}: {exc}"
            ) from exc
        measures.append(
            MeasureSpec(
                name=name,
                haz_type=haz_type,
                cost=cost,
                hazard_freq_cutoff=freq_cutoff,
                hazard_inten_imp=inten_imp,
                mdd_impact_a=mdd_a,
                mdd_impact_b=mdd_b,
                paa_impact_a=paa_a,
                paa_impact_b=paa_b,
            )
        )

    return measures


def _parse_discount_rate(wb: openpyxl.Workbook, path: Path) -> float:
    _require_sheet(wb, "discount", path)
    ws = wb["discount"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise EntityLoadError(f"Empty 'discount' sheet in {path.name}")

    header = _normalise_header(rows[0])
    rate_c = _require_col(header, "discount_rate", "discount", path)

    rates = [float(r[rate_c]) for r in rows[1:] if r[rate_c] is not None]
    if not rates:
        raise EntityLoadError(f"No discount_rate values found in 'discount' sheet of {path.name}")

    return float(np.mean(rates))


def _parse_ref_year(wb: openpyxl.Workbook, path: Path) -> int:
    if "names" not in wb.sheetnames:
        return 2020
    ws = wb["names"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return 2020

    header = _normalise_header(rows[0])
    item_c = _find_col(header, "item")
    name_c = _find_col(header, "name")
    if item_c is None or name_c is None:
        return 2020

    for row in rows[1:]:
        if row[item_c] is not None and str(row[item_c]).strip().lower() == "reference_year":
            val = row[name_c]
            if val is not None:
                try:
                    return int(val)
                except (TypeError, ValueError):
                    pass
    return 2020


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------


def _validate_specs(specs: list[ImpactFunctionSpec], path: Path) -> None:
    """Validate intensity monotonicity, unit consistency, and id uniqueness.

    mdd / paa monotonicity is intentionally not checked here: production
    CLIMADA entity files can have non-monotonic mdd curves (e.g. negative
    avoided-damage values that cross zero) which CLIMADA accepts. The
    ImpactFunctionRegistry enforces stricter monotonicity for JSON-registry
    specs; that check is not "applicable" for XLSX-sourced data.
    """
    triples: dict[tuple[str, str, int], str] = {}
    units_by_haz: dict[str, str] = {}

    for spec in specs:
        intensities = spec.intensity
        if len(intensities) >= 2 and not all(
            b >= a for a, b in zip(intensities, intensities[1:], strict=False)
        ):
            raise EntityLoadError(
                f"Impact function {spec.name!r} (haz_type={spec.haz_type!r}, id={spec.id}) "
                f"in {path.name}: 'intensity' must be non-decreasing, "
                f"got {list(intensities)!r}"
            )

        triple = (spec.haz_type, spec.exp_type, spec.id)
        if triple in triples:
            raise EntityLoadError(
                f"Duplicate impact-function id {spec.id} for "
                f"(haz_type={spec.haz_type!r}, exp_type={spec.exp_type!r}) in {path.name}; "
                f"already registered as {triples[triple]!r}"
            )
        triples[triple] = spec.name

        existing_unit = units_by_haz.get(spec.haz_type)
        if existing_unit is not None and existing_unit != spec.intensity_unit:
            raise EntityLoadError(
                f"Inconsistent intensity_unit for haz_type {spec.haz_type!r} in {path.name}: "
                f"got {spec.intensity_unit!r} but earlier function uses {existing_unit!r}"
            )
        units_by_haz[spec.haz_type] = spec.intensity_unit


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _normalise_header(row: tuple[Any, ...]) -> list[str]:
    return [str(c).strip().lower() if c is not None else "" for c in row]


def _require_sheet(wb: openpyxl.Workbook, sheet: str, path: Path) -> None:
    if sheet not in wb.sheetnames:
        raise EntityLoadError(f"Missing required sheet {sheet!r} in {path.name}")


def _require_col(header: list[str], col: str, sheet: str, path: Path) -> int:
    normalised = col.lower()
    try:
        return header.index(normalised)
    except ValueError:
        raise EntityLoadError(
            f"Missing required column {col!r} in sheet {sheet!r} of {path.name}"
        ) from None


def _find_col(header: list[str], col: str) -> int | None:
    normalised = col.lower()
    try:
        return header.index(normalised)
    except ValueError:
        return None


def _find_col_prefix(header: list[str], prefix: str) -> int | None:
    """Return the index of the first column whose name starts with *prefix* (case-insensitive)."""
    lower_prefix = prefix.lower()
    for i, col in enumerate(header):
        if col.startswith(lower_prefix):
            return i
    return None


@overload
def _opt_float(row: tuple[Any, ...], col: int | None, default: float) -> float: ...
@overload
def _opt_float(row: tuple[Any, ...], col: int | None, default: None = None) -> float | None: ...
def _opt_float(row: tuple[Any, ...], col: int | None, default: float | None = None) -> float | None:
    """Return float(row[col]) if col is not None and the cell is not None, else default."""
    if col is None or row[col] is None:
        return default
    return float(row[col])


def _opt_array(
    data: list[tuple[Any, ...]],
    col: int | None,
    field: str,
    path: Path,
) -> np.ndarray | None:
    """Build a float64 ndarray from an optional column, or return None if the column is absent."""
    if col is None:
        return None
    try:
        return np.array([float(r[col] or 0.0) for r in data], dtype=np.float64)
    except (TypeError, ValueError) as exc:
        raise EntityLoadError(
            f"Invalid {field!r} value in 'assets' sheet of {path.name}: {exc}"
        ) from exc
