"""Tests for the Excel report's Provenance sheet + BibTeX (#122).

The general/hazard/exposure/impact tabs read parquet files from
``DATA_TEMP_DIR`` and would require a CLIMADA fixture tree to exercise; the
provenance sheet is the only addition #122 makes, and it only needs a
seeded DuckDB row plus the small ``build_bibtex_snippet`` helper.
"""

from __future__ import annotations

from pathlib import Path

import duckdb
import pytest
from db import DB_PATH_ENV_VAR, MIGRATIONS_DIR, insert_scenario, run_migrations
from openpyxl import load_workbook
from provenance import short_sha
from report.report_handler import (
    ReportHandler,
    ReportParameters,
    build_bibtex_snippet,
)


class _ReportHandlerStub:
    """Light-weight stand-in: ``_generate_provenance_tab`` only reads
    ``self.report_parameters`` so we can exercise it without booting the
    real ReportHandler (which transitively imports CLIMADA via BaseHandler).
    """

    def __init__(self, report_parameters: ReportParameters) -> None:
        self.report_parameters = report_parameters

    _generate_provenance_tab = ReportHandler._generate_provenance_tab


class TestShortSha:
    @pytest.mark.parametrize(
        "value,expected",
        [
            ("abcdef0123456789", "abcdef01"),
            ("", "-"),
            (None, "-"),
        ],
    )
    def test_truncates_to_eight_chars(self, value, expected) -> None:
        assert short_sha(value) == expected


class TestBibtex:
    def test_includes_app_and_climada_versions(self) -> None:
        snippet = build_bibtex_snippet("2.0.0", "4.1.1")
        assert "@techreport{riskwise" in snippet
        assert "App v2.0.0" in snippet
        assert "CLIMADA v4.1.1" in snippet
        assert "UNU-EHS / GIZ" in snippet

    def test_handles_missing_versions(self) -> None:
        snippet = build_bibtex_snippet(None, None)
        assert "App v?, CLIMADA v?" in snippet


@pytest.fixture
def seeded_scenario(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> str:
    """Insert one scenario into a tmp DuckDB and return its id."""
    db_path = tmp_path / "report.db"
    monkeypatch.setenv(DB_PATH_ENV_VAR, str(db_path))
    conn = duckdb.connect(str(db_path))
    try:
        run_migrations(conn, migrations_dir=MIGRATIONS_DIR)
    finally:
        conn.close()
    insert_scenario(
        "scen-prov",
        {
            "country": "Egypt",
            "hazard_type": "flood",
            "scenario": "rcp85",
            "exposure_economic": "crops",
            "exposure_non_economic": "",
            "ref_year": 2024,
            "future_year": 2050,
            "annual_growth": 2.0,
            "is_era": True,
            "app_option": "era",
        },
        results={},
        provenance={
            "app_version": "2.0.0",
            "engine_version": "2.0.0",
            "climada_version": "4.1.1",
            "entity_data_sha256": "deadbeef" + "0" * 56,
            "hazard_data_sha256": "feedface" + "0" * 56,
            "country_config_sha256": "cafebabe" + "0" * 56,
            "config_version": "1",
            "random_seed": 42,
        },
        name="Egypt flood baseline",
    )
    return "scen-prov"


class TestProvenanceSheet:
    def test_sheet_contains_all_expected_fields(
        self, seeded_scenario: str, tmp_path: Path
    ) -> None:
        import xlsxwriter

        params = ReportParameters(
            scenario_id=seeded_scenario,
            country_name="Egypt",
            hazard="flood",
            scenario="rcp85",
            time_horizon="2024 - 2050",
        )
        handler = _ReportHandlerStub(params)
        out_path = tmp_path / "report.xlsx"
        wb = xlsxwriter.Workbook(filename=str(out_path))
        handler._generate_provenance_tab(wb)
        wb.close()

        loaded = load_workbook(out_path)
        assert "Provenance" in loaded.sheetnames
        ws = loaded["Provenance"]
        flat = " ".join(
            str(cell.value)
            for row in ws.iter_rows()
            for cell in row
            if cell.value is not None
        )
        # 8-char prefix shown for each SHA, never the full 64 chars.
        assert "deadbeef" in flat
        assert "deadbeef" + "0" * 56 not in flat
        assert "feedface" in flat
        assert "cafebabe" in flat
        # Provenance values surfaced.
        assert "2.0.0" in flat
        assert "4.1.1" in flat
        assert "42" in flat
        # Caveat note + BibTeX present.
        assert "Cross-platform" in flat or "BLAS" in flat
        assert "@techreport{riskwise" in flat

    def test_no_sheet_when_scenario_unknown(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        import xlsxwriter

        # Point at a fresh migrated DB so the get_scenario lookup returns None
        # cleanly instead of hitting the developer's local workspace DB (or
        # a stale one missing the 0006 migration).
        db_path = tmp_path / "empty.db"
        monkeypatch.setenv(DB_PATH_ENV_VAR, str(db_path))
        conn = duckdb.connect(str(db_path))
        try:
            run_migrations(conn, migrations_dir=MIGRATIONS_DIR)
        finally:
            conn.close()

        params = ReportParameters(scenario_id="does-not-exist")
        handler = _ReportHandlerStub(params)
        out_path = tmp_path / "empty.xlsx"
        wb = xlsxwriter.Workbook(filename=str(out_path))
        # Stub sheet so xlsxwriter can close (it refuses empty workbooks).
        wb.add_worksheet("Sentinel")
        handler._generate_provenance_tab(wb)
        wb.close()
        loaded = load_workbook(out_path)
        assert "Provenance" not in loaded.sheetnames
        assert "Sentinel" in loaded.sheetnames
