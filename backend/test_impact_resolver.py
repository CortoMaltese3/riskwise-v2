"""Tests for :mod:`backend.impact.resolver`.

The resolver delegates to the entity XLSX loader the engine already uses, so
these tests focus on filename derivation, error translation, and the
disambiguation rule when a workbook contains more than one function for a
single hazard. Fixtures are built in-memory via ``openpyxl`` (the same pattern
as ``tests/unit/engine/loaders/test_xlsx.py``) so the suite stays
hermetic — no CLIMADA, no shipped data.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from backend.impact.resolver import (
    ImpactFunctionNotFoundError,
    get_active_impact_function,
)


def _write_entity_workbook(
    path: Path,
    impf_rows: list[tuple[int, float, float, float, str, str, str]],
    asset_impf_id: int = 1,
) -> None:
    """Write a minimal entity workbook with the given impact-function curve rows.

    Each ``impf_rows`` entry is ``(id, intensity, mdd, paa, peril_id, unit, name)``.
    """
    wb = openpyxl.Workbook()
    wb.active.title = "assets"

    ws_assets = wb["assets"]
    ws_assets.append(
        ["latitude", "longitude", "value", "value_unit", "deductible", "cover", "impf_"]
    )
    ws_assets.append([10.0, 20.0, 1_000_000.0, "USD", 0.0, 1_000_000.0, asset_impf_id])
    ws_assets.append([11.0, 21.0, 2_000_000.0, "USD", 0.0, 2_000_000.0, asset_impf_id])

    ws_impf = wb.create_sheet("impact_functions")
    ws_impf.append(
        ["impact_fun_id", "intensity", "mdd", "paa", "peril_id", "intensity_unit", "name"]
    )
    for row in impf_rows:
        ws_impf.append(list(row))

    ws_meas = wb.create_sheet("measures")
    ws_meas.append(["name", "peril_ID", "cost"])
    ws_meas.append(["M", "FL", 100.0])

    ws_disc = wb.create_sheet("discount")
    ws_disc.append(["year", "discount_rate"])
    ws_disc.append([2024, 0.02])

    wb.save(path)


def _curve(impf_id: int, name: str, haz: str = "FL", unit: str = "m") -> list[tuple]:
    return [
        (impf_id, 0.0, 0.0, 1.0, haz, unit, name),
        (impf_id, 1.0, 0.5, 1.0, haz, unit, name),
        (impf_id, 2.0, 1.0, 1.0, haz, unit, name),
    ]


class TestCanonicalPath:
    def test_resolves_era_canonical_filename(self, tmp_path: Path, monkeypatch) -> None:
        monkeypatch.setattr("backend.impact.resolver.DATA_ENTITIES_DIR", tmp_path)
        path = tmp_path / "entity_TODAY_EGY_FL_diarrhea_patients.xlsx"
        _write_entity_workbook(path, _curve(105, "Diarrhoea patients"))

        spec = get_active_impact_function("Egypt", "flood", "diarrhea_patients")

        assert spec.id == 105
        assert spec.name == "Diarrhoea patients"
        assert spec.haz_type == "FL"
        assert spec.intensity_unit == "m"
        assert spec.intensity == (0.0, 1.0, 2.0)
        assert spec.mdd == (0.0, 0.5, 1.0)

    def test_uses_hazard_alias_mapping(self, tmp_path: Path, monkeypatch) -> None:
        # ``flash_flood`` and ``flood`` both map to the ``FL`` peril code.
        monkeypatch.setattr("backend.impact.resolver.DATA_ENTITIES_DIR", tmp_path)
        path = tmp_path / "entity_TODAY_THA_FL_crops.xlsx"
        _write_entity_workbook(path, _curve(50, "Crops"))

        spec = get_active_impact_function("Thailand", "flash_flood", "crops")
        assert spec.haz_type == "FL"


class TestCustomPath:
    def test_custom_entity_file_overrides_canonical_lookup(
        self, tmp_path: Path, monkeypatch
    ) -> None:
        monkeypatch.setattr("backend.impact.resolver.DATA_ENTITIES_DIR", tmp_path)
        custom = tmp_path / "custom_upload.xlsx"
        _write_entity_workbook(custom, _curve(999, "Custom function"))

        spec = get_active_impact_function(
            "Egypt", "flood", "anything", entity_file="custom_upload.xlsx"
        )
        assert spec.id == 999
        assert spec.name == "Custom function"

    def test_custom_entity_file_accepts_absolute_path(self, tmp_path: Path, monkeypatch) -> None:
        elsewhere = tmp_path / "outside"
        elsewhere.mkdir()
        path = elsewhere / "weird_name.xlsx"
        _write_entity_workbook(path, _curve(42, "Elsewhere"))

        monkeypatch.setattr("backend.impact.resolver.DATA_ENTITIES_DIR", tmp_path / "entities")
        spec = get_active_impact_function("Egypt", "flood", "ignored", entity_file=str(path))
        assert spec.id == 42


class TestFailures:
    def test_unknown_hazard_raises_not_found(self, tmp_path: Path, monkeypatch) -> None:
        monkeypatch.setattr("backend.impact.resolver.DATA_ENTITIES_DIR", tmp_path)
        with pytest.raises(ImpactFunctionNotFoundError, match="hazard"):
            get_active_impact_function("Egypt", "asteroid_strike", "crops")

    def test_unknown_country_raises_not_found(self, tmp_path: Path, monkeypatch) -> None:
        monkeypatch.setattr("backend.impact.resolver.DATA_ENTITIES_DIR", tmp_path)
        with pytest.raises(ImpactFunctionNotFoundError, match="country"):
            get_active_impact_function("Atlantis", "flood", "crops")

    def test_missing_xlsx_raises_not_found(self, tmp_path: Path, monkeypatch) -> None:
        monkeypatch.setattr("backend.impact.resolver.DATA_ENTITIES_DIR", tmp_path)
        with pytest.raises(ImpactFunctionNotFoundError, match="not found"):
            get_active_impact_function("Egypt", "flood", "crops")

    def test_hazard_with_no_matching_spec_raises_not_found(
        self, tmp_path: Path, monkeypatch
    ) -> None:
        monkeypatch.setattr("backend.impact.resolver.DATA_ENTITIES_DIR", tmp_path)
        path = tmp_path / "only_drought.xlsx"
        # Workbook contains a drought function but we ask for flood.
        _write_entity_workbook(path, _curve(7, "Drought", haz="D", unit="SPI"))
        with pytest.raises(ImpactFunctionNotFoundError, match="haz_type='FL'"):
            get_active_impact_function("Egypt", "flood", "ignored", entity_file="only_drought.xlsx")


class TestEngineParity:
    """The viewer must read the same XLSX the engine consumes at run time."""

    def test_resolver_and_engine_loader_return_the_same_spec(
        self, tmp_path: Path, monkeypatch
    ) -> None:
        # The engine path: ``load_entity_xlsx`` produces the ``impfset_specs``
        # list ``RunScenario._execute`` hands to ``calculate_impact``. The
        # viewer must surface a spec drawn from the same list — otherwise the
        # panel can lie about what the engine will actually apply.
        from backend.engine.loaders.xlsx import load_entity_xlsx

        monkeypatch.setattr("backend.impact.resolver.DATA_ENTITIES_DIR", tmp_path)
        path = tmp_path / "entity_TODAY_EGY_FL_crops.xlsx"
        _write_entity_workbook(path, _curve(12, "Crops"))

        engine_specs = load_entity_xlsx(path).impfset_specs
        viewer_spec = get_active_impact_function("Egypt", "flood", "crops")

        assert viewer_spec in engine_specs
        assert viewer_spec.id == 12
        assert viewer_spec.name == "Crops"


class TestMultiSpecDisambiguation:
    def test_picks_spec_matching_most_common_impf_id(self, tmp_path: Path, monkeypatch) -> None:
        monkeypatch.setattr("backend.impact.resolver.DATA_ENTITIES_DIR", tmp_path)
        path = tmp_path / "multi.xlsx"
        rows = _curve(1, "First") + _curve(2, "Second")
        # Exposures point at impf_id=2, so the resolver should return the
        # second spec even though both share haz_type=FL.
        _write_entity_workbook(path, rows, asset_impf_id=2)

        spec = get_active_impact_function("Egypt", "flood", "ignored", entity_file="multi.xlsx")
        assert spec.id == 2
        assert spec.name == "Second"
