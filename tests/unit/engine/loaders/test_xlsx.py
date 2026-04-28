"""Tests for :func:`backend.engine.loaders.xlsx.load_entity_xlsx`.

Covers:

* Loader-output structure and field types on the two reference fixtures.
* Parity against CLIMADA's ``Entity.from_excel`` on the same fixtures
  (skipped when CLIMADA is not importable or is only stubbed).
* Validation-failure paths (non-monotonic intensity, duplicate impf id,
  missing required sheet).
"""

from __future__ import annotations

import re
from pathlib import Path

import numpy as np
import openpyxl
import pytest

from backend.engine.loaders import EntityLoadError, load_entity_xlsx
from backend.engine.types import EntityBundle, ExposureArrays, ImpactFunctionSpec, MeasureSpec

FIXTURES = Path(__file__).parents[3] / "fixtures" / "entities"
EGY = FIXTURES / "egy_economic_present.xlsx"
THA = FIXTURES / "tha_economic_present.xlsx"


@pytest.fixture
def egy_path() -> Path:
    if not EGY.is_file():
        pytest.skip(f"missing fixture: {EGY}")
    return EGY


@pytest.fixture
def tha_path() -> Path:
    if not THA.is_file():
        pytest.skip(f"missing fixture: {THA}")
    return THA


# ---------------------------------------------------------------------------
# Structural checks
# ---------------------------------------------------------------------------


def test_load_egy_returns_entity_bundle(egy_path: Path) -> None:
    bundle = load_entity_xlsx(egy_path)

    assert isinstance(bundle, EntityBundle)
    assert isinstance(bundle.exposures, ExposureArrays)
    assert bundle.exposures.values.shape == (50,)
    assert bundle.exposures.centroid_idx.shape == (50,)
    assert bundle.exposures.impf_id.shape == (50,)
    assert bundle.exposures.lat.shape == (50,)
    assert bundle.exposures.lon.shape == (50,)
    assert bundle.exposures.deductible is not None
    assert bundle.exposures.cover is not None
    assert np.isfinite(bundle.exposures.values).all()
    assert np.isfinite(bundle.exposures.lat).all()
    assert np.isfinite(bundle.exposures.lon).all()


def test_load_tha_returns_entity_bundle(tha_path: Path) -> None:
    bundle = load_entity_xlsx(tha_path)

    assert isinstance(bundle, EntityBundle)
    assert bundle.exposures.values.shape == (50,)
    assert bundle.exposures.centroid_idx.shape == (50,)
    assert np.isfinite(bundle.exposures.values).all()


def test_centroid_idx_is_sequential_arange(egy_path: Path) -> None:
    bundle = load_entity_xlsx(egy_path)
    np.testing.assert_array_equal(bundle.exposures.centroid_idx, np.arange(50))


def test_impfset_specs_populated(egy_path: Path) -> None:
    bundle = load_entity_xlsx(egy_path)
    assert len(bundle.impfset_specs) > 0
    for spec in bundle.impfset_specs:
        assert isinstance(spec, ImpactFunctionSpec)
        assert len(spec.intensity) >= 2
        assert len(spec.intensity) == len(spec.mdd) == len(spec.paa)
        assert spec.haz_type != ""
        assert spec.intensity_unit != ""


def test_intensity_is_non_decreasing(egy_path: Path) -> None:
    bundle = load_entity_xlsx(egy_path)
    for spec in bundle.impfset_specs:
        diffs = np.diff(np.array(spec.intensity))
        assert (diffs >= 0).all(), f"Non-decreasing violated for {spec.name!r}"


def test_measures_populated(egy_path: Path) -> None:
    bundle = load_entity_xlsx(egy_path)
    assert len(bundle.measures) > 0
    for m in bundle.measures:
        assert isinstance(m, MeasureSpec)
        assert m.name != ""
        assert m.haz_type != ""


def test_discount_rate_is_positive_float(egy_path: Path) -> None:
    bundle = load_entity_xlsx(egy_path)
    assert isinstance(bundle.discount_rate, float)
    assert bundle.discount_rate > 0


def test_ref_year_is_integer(egy_path: Path) -> None:
    bundle = load_entity_xlsx(egy_path)
    assert isinstance(bundle.ref_year, int)
    assert 2000 <= bundle.ref_year <= 2100


# ---------------------------------------------------------------------------
# Parity vs CLIMADA's Entity.from_excel
# ---------------------------------------------------------------------------


def _climada_entity(path: Path):
    entity_mod = pytest.importorskip("climada.entity")
    if not hasattr(entity_mod.Entity, "from_excel"):
        pytest.skip("climada is stubbed — real package not installed")
    return entity_mod.Entity.from_excel(str(path))


def _climada_impf_col(gdf) -> str:
    """Find the impf_* column name in a CLIMADA Exposures GeoDataFrame."""
    for col in gdf.columns:
        if col.startswith("impf_"):
            return col
    raise AssertionError(f"No impf_* column found in GDF columns: {list(gdf.columns)}")


@pytest.mark.parametrize(
    ("fixture_name", "path"),
    [("egy", EGY), ("tha", THA)],
)
def test_climada_parity_exposures(fixture_name: str, path: Path) -> None:
    if not path.is_file():
        pytest.skip(f"missing fixture: {path}")
    entity = _climada_entity(path)
    bundle = load_entity_xlsx(path)

    gdf = entity.exposures.gdf
    impf_col = _climada_impf_col(gdf)

    np.testing.assert_allclose(
        bundle.exposures.values,
        gdf["value"].to_numpy(dtype=np.float64),
        atol=1e-9,
        rtol=0,
        err_msg=f"{fixture_name}: values mismatch",
    )
    np.testing.assert_allclose(
        bundle.exposures.lat,
        gdf["latitude"].to_numpy(dtype=np.float64),
        atol=1e-9,
        rtol=0,
        err_msg=f"{fixture_name}: lat mismatch",
    )
    np.testing.assert_allclose(
        bundle.exposures.lon,
        gdf["longitude"].to_numpy(dtype=np.float64),
        atol=1e-9,
        rtol=0,
        err_msg=f"{fixture_name}: lon mismatch",
    )
    np.testing.assert_allclose(
        bundle.exposures.impf_id.astype(np.float64),
        gdf[impf_col].to_numpy(dtype=np.float64),
        atol=1e-9,
        rtol=0,
        err_msg=f"{fixture_name}: impf_id mismatch",
    )


@pytest.mark.parametrize(
    ("fixture_name", "path"),
    [("egy", EGY), ("tha", THA)],
)
def test_climada_parity_impact_functions(fixture_name: str, path: Path) -> None:
    if not path.is_file():
        pytest.skip(f"missing fixture: {path}")
    entity = _climada_entity(path)
    bundle = load_entity_xlsx(path)

    impfset = entity.impact_funcs
    for spec in bundle.impfset_specs:
        climada_func = impfset.get_func(haz_type=spec.haz_type, fun_id=spec.id)
        np.testing.assert_allclose(
            np.array(spec.intensity),
            climada_func.intensity,
            atol=1e-12,
            rtol=0,
            err_msg=f"{fixture_name}: intensity mismatch for {spec.name!r}",
        )
        np.testing.assert_allclose(
            np.array(spec.mdd),
            climada_func.mdd,
            atol=1e-12,
            rtol=0,
            err_msg=f"{fixture_name}: mdd mismatch for {spec.name!r}",
        )
        np.testing.assert_allclose(
            np.array(spec.paa),
            climada_func.paa,
            atol=1e-12,
            rtol=0,
            err_msg=f"{fixture_name}: paa mismatch for {spec.name!r}",
        )


@pytest.mark.parametrize(
    ("fixture_name", "path"),
    [("egy", EGY), ("tha", THA)],
)
def test_climada_parity_measures(fixture_name: str, path: Path) -> None:
    if not path.is_file():
        pytest.skip(f"missing fixture: {path}")
    entity = _climada_entity(path)
    bundle = load_entity_xlsx(path)

    measure_set = entity.measures
    climada_names = {
        m.name for haz_measures in measure_set.get_measure().values() for m in haz_measures
    }
    our_names = {m.name for m in bundle.measures}
    assert our_names == climada_names, f"{fixture_name}: measure name sets differ"


# ---------------------------------------------------------------------------
# Validation failure paths
# ---------------------------------------------------------------------------


def test_missing_file_raises_entity_load_error(tmp_path: Path) -> None:
    missing = tmp_path / "does_not_exist.xlsx"
    with pytest.raises(EntityLoadError, match=re.escape(str(missing))):
        load_entity_xlsx(missing)


def test_missing_sheet_raises_entity_load_error(tmp_path: Path) -> None:
    bad = tmp_path / "no_assets.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "impact_functions"
    wb.save(bad)
    with pytest.raises(EntityLoadError, match="assets"):
        load_entity_xlsx(bad)


def _make_minimal_xlsx(tmp_path: Path, filename: str) -> tuple[Path, openpyxl.Workbook]:
    """Return a path and an open workbook for a minimal valid entity XLSX."""
    path = tmp_path / filename
    wb = openpyxl.Workbook()
    wb.active.title = "assets"

    ws_assets = wb["assets"]
    ws_assets.append(
        ["latitude", "longitude", "value", "value_unit", "deductible", "cover", "impf_"]
    )
    ws_assets.append([10.0, 20.0, 1_000_000.0, "USD", 0.0, 1_000_000.0, 1])
    ws_assets.append([11.0, 21.0, 2_000_000.0, "USD", 0.0, 2_000_000.0, 1])

    ws_impf = wb.create_sheet("impact_functions")
    ws_impf.append(
        ["impact_fun_id", "intensity", "mdd", "paa", "peril_id", "intensity_unit", "name"]
    )
    ws_impf.append([1, 0.0, 0.0, 1.0, "FL", "m", "Test"])
    ws_impf.append([1, 1.0, 0.5, 1.0, "FL", "m", "Test"])
    ws_impf.append([1, 2.0, 1.0, 1.0, "FL", "m", "Test"])

    ws_meas = wb.create_sheet("measures")
    ws_meas.append(["name", "peril_ID", "cost"])
    ws_meas.append(["TestMeasure", "FL", 100_000.0])

    ws_disc = wb.create_sheet("discount")
    ws_disc.append(["year", "discount_rate"])
    ws_disc.append([2024, 0.02])
    ws_disc.append([2025, 0.03])

    ws_names = wb.create_sheet("names")
    ws_names.append(["Item", "ID", "name"])
    ws_names.append(["reference_year", 0, 2024])

    return path, wb


def test_valid_minimal_workbook_loads(tmp_path: Path) -> None:
    path, wb = _make_minimal_xlsx(tmp_path, "valid.xlsx")
    wb.save(path)
    bundle = load_entity_xlsx(path)
    assert bundle.exposures.values.shape == (2,)
    assert bundle.ref_year == 2024
    assert abs(bundle.discount_rate - 0.025) < 1e-12


def test_non_monotonic_intensity_raises_entity_load_error(tmp_path: Path) -> None:
    path = tmp_path / "bad_intensity.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "assets"
    ws_assets = wb["assets"]
    ws_assets.append(
        ["latitude", "longitude", "value", "value_unit", "deductible", "cover", "impf_"]
    )
    ws_assets.append([10.0, 20.0, 1_000_000.0, "USD", 0.0, 1_000_000.0, 1])

    ws_impf = wb.create_sheet("impact_functions")
    ws_impf.append(
        ["impact_fun_id", "intensity", "mdd", "paa", "peril_id", "intensity_unit", "name"]
    )
    ws_impf.append([1, 2.0, 0.0, 1.0, "FL", "m", "Test"])  # start at 2
    ws_impf.append([1, 1.0, 0.5, 1.0, "FL", "m", "Test"])  # drops to 1 — non-monotonic
    ws_impf.append([1, 3.0, 1.0, 1.0, "FL", "m", "Test"])

    ws_meas = wb.create_sheet("measures")
    ws_meas.append(["name", "peril_ID", "cost"])
    ws_meas.append(["M", "FL", 100.0])

    ws_disc = wb.create_sheet("discount")
    ws_disc.append(["year", "discount_rate"])
    ws_disc.append([2024, 0.02])

    wb.save(path)
    with pytest.raises(EntityLoadError, match="intensity"):
        load_entity_xlsx(path)


def test_duplicate_impf_id_raises_entity_load_error(tmp_path: Path) -> None:
    path, wb = _make_minimal_xlsx(tmp_path, "dup_id.xlsx")
    ws_impf = wb["impact_functions"]
    # Two rows with same (id, haz, name) but different unit form separate groups,
    # both producing an ImpactFunctionSpec with the same (haz_type, exp_type, id)
    # triple → triggers the duplicate-id validation.
    ws_impf.append([1, 3.0, 1.0, 1.0, "FL", "km", "Test"])
    wb.save(path)
    with pytest.raises(EntityLoadError, match=r"[Dd]uplicate"):
        load_entity_xlsx(path)
