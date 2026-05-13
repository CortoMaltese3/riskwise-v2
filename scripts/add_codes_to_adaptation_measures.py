"""Add a ``code`` column to ``requirements/adaptation_measures.xlsx``.

Run once to backfill catalog codes that align engine ``measure_name``
(short codes the entity xlsxs ship — "GR", "TP", ...) with the catalog's
i18n key (``adaptation_measures_green_roofs`` ...). Codes that are
ambiguous against the entity xlsxs are left blank pending data-owner
clarification — see issue #429.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl

# Mapping is from catalog i18n-key tail (the part after the
# ``adaptation_measures_`` prefix) to engine code. Only the 16 unambiguous
# codes enumerated in #429 are populated; the remaining rows ship a blank
# code and fall back to the raw engine name in the chart.
TAIL_TO_CODE: dict[str, str] = {
    "green_roofs": "GR",
    "trees_planting": "TP",
    "green_spaces": "GS",
    "early_warning_system": "EWS",
    "climate_smart_agriculture": "CSA",
    "green_building_codes": "GBC",
    "management_of_protected_environmental_areas": "MPA",
    "retention_reservoirs": "RR",
    "dredging_of_canals": "DC",
    "infiltration_ponds": "IP",
    "recharging_wells": "RW",
    "small_dams": "SD",
    "retention_furrows": "RF",
    "agricultural_zoning": "AZ",
    "water_tolerant_crops": "WTC",
    "flood_index_insurance": "FI",
}

PREFIX = "adaptation_measures_"


def code_for_name(name: str) -> str:
    """Return the engine code for a catalog i18n key, or empty string."""
    if not isinstance(name, str) or not name.startswith(PREFIX):
        return ""
    return TAIL_TO_CODE.get(name[len(PREFIX) :], "")


def main(xlsx_path: Path) -> None:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["measures"]

    header = [cell.value for cell in ws[1]]
    if "code" in header:
        name_idx = header.index("name") + 1
        code_idx = header.index("code") + 1
        for row in ws.iter_rows(min_row=2, values_only=False):
            name_val = row[name_idx - 1].value
            row[code_idx - 1].value = code_for_name(name_val)
    else:
        code_idx = len(header) + 1
        ws.cell(row=1, column=code_idx, value="code")
        name_idx = header.index("name") + 1
        for row_num in range(2, ws.max_row + 1):
            name_val = ws.cell(row=row_num, column=name_idx).value
            ws.cell(row=row_num, column=code_idx, value=code_for_name(name_val))

    wb.save(xlsx_path)


if __name__ == "__main__":
    repo_root = Path(__file__).resolve().parents[1]
    main(repo_root / "requirements" / "adaptation_measures.xlsx")
